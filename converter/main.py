#/usr/bin/python

import argparse
import csv
import datetime
import logging
import os

from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
from converter.ordered_set import OrderedSet

COMBINED_XLS_FILE='/Users/alan/Library/Mobile Documents/com~apple~CloudDocs/Drumcree/Vestry/Envelopes/empty_book.xlsx'
CHURCH_SUITE_XLS_FILE='/Users/alan/Library/Mobile Documents/com~apple~CloudDocs/Drumcree/Vestry/Envelopes/churchsuite.xlsx'
CHURCH_SUITE_CSV_FILE='/Users/alan/Library/Mobile Documents/com~apple~CloudDocs/Drumcree/Vestry/Envelopes/churchsuite.csv'

# create logger
logger = logging.getLogger('Finance Converter')
logger.setLevel(logging.DEBUG)

# create console handler and set level to debug
ch = logging.StreamHandler()
ch.setLevel(logging.DEBUG)

# create formatter
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# add formatter to ch
ch.setFormatter(formatter)

# add ch to logger
logger.addHandler(ch)


def load_spreadsheet(location):
    logger.debug(location)
    spreadsheet = load_workbook(location, data_only=True)
    logger.debug('Spreadsheet Names: {}'.format(spreadsheet.sheetnames))
    return spreadsheet

def iter_rows(row, n):  #produce the list of items in the particular row
    yield [cell.value for cell in row]

def print_sheet(ws):
    for row in ws.iter_rows():
        yield [str((cell.value)).strip() for cell in row]

def combine_sheets(spreadsheet, date_arg):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "All Envelops"

    envelop_sheets = ['1-33', '34-66', '67-100', '101-133', '134-166', '167-200', '201-233']

    first_sheet = True

    rows_to_ignore = [1,]
    columns_to_delete = ['AA']

    for sheet in envelop_sheets:
        for row in spreadsheet[sheet].rows:
            current_row = row[0].row
            current_value = row[0].value

            # logger.debug('Current Sheet - {}, Current Row - {}'.format(sheet, current_row))
            # logger.debug('Envelope Number - {}'.format(current_value))

            if first_sheet or (current_row not in rows_to_ignore and current_value != 'Total'):
                n = 'A' + str(row) + ':' + ('Z' + str(row))
                   
                list_to_append = list(iter_rows(row,n))

                #logger.debug('List to append - {}'.format(list_to_append))
                for items in list_to_append:
                    # logger.debug('Item 1: {}'.format(items[0]))
                    ws1.append(items)
    
                if current_row == 1:
                    first_sheet = False

    logger.debug('Columns to delete: {}'.format(columns_to_delete))
    for col in OrderedSet(columns_to_delete):
        cidx = column_index_from_string(col)
        ws1.delete_cols(cidx)

    colnames = date_arg
    col_indices = {}
    cols_indices_to_delete = {}

    for column in ws1.columns:
        currentCell = column[0]
        # logger.debug('Column Name: {}'.format(currentCell.value))
        if currentCell.value in colnames:
            col_indices[currentCell.column] = currentCell.value
        elif currentCell.column != 'A':
            cols_indices_to_delete[currentCell.column] = currentCell.value

    for col, val in cols_indices_to_delete.items():
        for c in ws1.columns:
            currentCell = c[0]
            if currentCell.value == val:
                cidx = column_index_from_string(currentCell.column)
                ws1.delete_cols(cidx)

    wb.save(filename=COMBINED_XLS_FILE)

    return wb

def convert_to_churchsuite_format(workbook):
    churchsuite_wb = Workbook()
    ws1 = churchsuite_wb.active
    ws1.title = "Lodgements"

    ws1.append(['bank_reference','amount','date','fund','method'])

    active_sheet = workbook["All Envelops"]
    
    for column in active_sheet.columns:
        if any(cell.value for cell in column) and column[0].value != "ENVELOPE NO":
            col_idx = column[0].col_idx
            current_date = column[0].value
        
            for row in column:
                if row.row !=1:
                    giver_envelop_num = active_sheet.cell(row.row, 1).value
                    amount_of_money = active_sheet.cell(row.row, col_idx).value
                    fund = 'Current Account'
                    method = 'Cash'
                    
                    if amount_of_money != None:
                        try:
                            d = current_date.strftime('%d-%m-%Y')
                        except:
                            d = "31-12-2018"

                        entry = (giver_envelop_num, amount_of_money, d, fund, method)
                        ws1.append(entry)

    churchsuite_wb.save(filename=CHURCH_SUITE_XLS_FILE)
    # sheet_value_arr = []
    with open(CHURCH_SUITE_CSV_FILE, 'w') as f:
        c = csv.writer(f)
        for row in ws1.rows:
#             sheet_value_arr.append([cell.value for cell in row])
#             for _ in ws1.rows:
            c.writerow([cell.value for cell in row])

        f.close()

def lookup_value(input_value, finance_spreadsheet):
    ws1 = map_spreadsheet.active

    for row in ws1.rows:
        for cell in row:
            if cell.col_idx == 1 and cell.value==input_value:
                return ws1.cell(cell.row, 2).value

def swap_ids(map_spreadsheet, finance_spreadsheet):
    ws_f = finance_spreadsheet.active

    for row in ws_f:
        for cell in row:
            if cell.col_idx == 1:
                v = ws_f.cell(cell.row, 1).value
                l_value = lookup_value(v, finance_spreadsheet)

                if l_value:
                    # print('Env number: {},  Online ID: {}'.format(v, l_value))
                    ws_f.cell(cell.row, 1).value = l_value


if __name__ == "__main__":
    
    # Lets Define and Parse the args
    parser = argparse.ArgumentParser(description="""Convert Finance Spreadsheet to Church Suite CSV
    e.g. --votes <PATH A>,<PATH B>,<PATH C> --candidates /Users/alan/votes/candidates.txt
    """)
    parser.add_argument('--input', type=str, help='The input Finance Spreadsheet. e.g. ')
    parser.add_argument('--date', type=str, nargs='*', help='The week to convert')
    parser.add_argument('--map', type=str, help='Spreadsheet of envelope numbers to online ID')
    #parser.add_argument('--out', type=str, help='The name of the output file')

    args = vars(parser.parse_args())

    # Lets extract the args into variables
    input_arg = os.path.expanduser(args['input'])
    try:
        date_arg = [datetime.datetime.strptime(a, '%d/%m/%Y') for a in args['date']]
    except:
        date_arg = args['date']
    
    # Now Lets log those for debugging
    logger.debug('Loading spreadsheet from: {}'.format(input_arg))
    logger.debug('Given Date: {}'.format(date_arg))
    
    # The meaty part
    finance_spreadsheet = load_spreadsheet(input_arg)
    wb = combine_sheets(finance_spreadsheet, date_arg)
    
    map_spreadsheet = load_spreadsheet(args['map'])
    swap_ids(map_spreadsheet, wb)
    wb.save(filename=COMBINED_XLS_FILE)

    convert_to_churchsuite_format(wb)
